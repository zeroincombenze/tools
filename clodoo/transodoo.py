#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# Copyright SHS-AV s.r.l. <http://ww.zeroincombenze.it>)
#
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
#
#    All Rights Reserved
#
"""Translate Odoo name to migrate from version to another
Structure:
[pymodel]:       Odoo model name
[pymodel][ttype]: may be:
    name:     symbolic name of a field (deprecated)
    field:    field name to translate
    action:   action/function to translate
    xref:     external reference (model is ir.model.data)
    model:    model name to translate (model is ir.model)
    module:   module name to translate (model is ir.module.module)
    merge:    module name or model merged with
    value:    value of field to translate (name is the field name)
    valuetnl: field ha translation ("1")

[pymodel][ttype][hash]           hash entry with ver.name list
[pymodel][ttype][hash][ver.name] value for specific version
[pymodel][ttype][ver.name]       specific name entry for name.ver -> hash

the ttype 'value' has a more level for every field name:
[pymodel]['value'][fldname][hash]
[pymodel]['value'][fldname][hash][ver.name]
[pymodel]['value'][fldname][ver.name]
"""
from __future__ import print_function, unicode_literals

import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from past.builtins import basestring

from python_plus import bstrings, unicodes

try:
    import jsonlib
except ImportError:
    pass
try:
    from z0lib.z0lib import z0lib
except ImportError:
    try:
        from z0lib import z0lib
    except ImportError:
        import z0lib

__version__ = "2.0.13"
VERSIONS = [
    "6.1",
    "7.0",
    "8.0",
    "9.0",
    "10.0",
    "11.0",
    "12.0",
    "13.0",
    "14.0",
    "15.0",
    "16.0",
]
ALL_VERSIONS = [x for x in VERSIONS]
for org in ("zero", "powerp", "librerp"):
    for ver in VERSIONS:
        if org == "librerp" and ver not in ("6.1", "12.0"):
            continue
        elif org == "powerp" and int(ver.split(".")[0]) < 12:
            continue
        ALL_VERSIONS.append("%s%s" % (org, ver.split(".")[0]))
CVT_ACC_TYPE_OLD_NEW = {
    "Bank": "Bank and Cash",
    "Cash": "Bank and Cash",
    "Check": "Credit Card",
    "Asset": "Current Assets",
    "Liability": "Current Liabilities",
    "Tax": "Current Liabilities",
}
CVT_ACC_TYPE_NEW_OLD = {
    "Bank and Cash": "Bank",
    "Credit Card": "Check",
    "Current Assets": "Asset",
    "Non-current Assets": "Asset",
    "Fixed Asset": "Asset",
    "Current Liabilities": "Liability",
    "Non-current Liabilities": "Liability",
    "Other Income": "Income",
    "Depreciation": "Expense",
    "Cost of Revenue": "Expense",
    "Prepayments": "Expense",
    "Current Year Earnings": "Expense",
}


def natstr(s):
    if sys.version_info[0] == 2:
        return bstrings(s)
    return unicodes(s)


def get_pymodel(model, ttype=None):
    return (
        {"xref": "ir.model.data", "model": "ir.model", "module": "ir.module.module"}
        .get(ttype, model or "res.groups")
        .lower()
    )


def get_ver_name(name, cur_ver):
    if name:
        return "%s~%s" % (cur_ver, name)
    return name


def is_hash(name):
    if re.match(r"^[a-zA-Z0-9_]+[a-zA-Z0-9_.-]*[0-9]+~", name):
        return False
    return True


def set_hash(ttype, name, ver_names):
    if ttype == "name":
        return name.upper()
    key = name if ttype in ("value", "valuetnl") else ""
    for item in ver_names:
        if key:
            key = "%s|%s" % (key, item)
        else:
            key = item
    if ttype in ("value", "valuetnl"):
        return key
    return key.upper()


def build_alias_struct(mindroot, model, ttype, fld_name=False):
    pymodel = get_pymodel(model, ttype=ttype)
    mindroot[pymodel] = mindroot.get(pymodel, {})
    mindroot[pymodel][ttype] = mindroot[pymodel].get(ttype, {})
    if ttype in ("value", "valuetnl") and fld_name:
        mindroot[pymodel][ttype][fld_name] = mindroot[pymodel][ttype].get(fld_name, {})
    return mindroot


def link_versioned_name(mindroot, model, hashed, ttype, src_name, ver, fld_name=False):
    pymodel = get_pymodel(model, ttype=ttype)
    ver_name = get_ver_name(src_name, ver)
    if ttype in ("value", "valuetnl"):
        if src_name and src_name.startswith("${") and src_name.endswith("}"):
            mindroot[pymodel][ttype][fld_name] = src_name
            item = None
        else:
            item = mindroot[pymodel][ttype][fld_name]
    else:
        item = mindroot[pymodel][ttype]
    if item is not None:
        item[hashed] = item.get(hashed, {})
        if ver_name:
            tver = ver_name
            while tver:
                tver = previous_ver_name(tver, ver_name)
                if tver in item and hashed not in item[tver]:
                    item[ver_name] = item[tver]
            if ver_name in item:
                if not isinstance(item[ver_name], list):
                    item[ver_name] = [item[ver_name]]
                if hashed not in item[ver_name]:
                    item[ver_name].append(hashed)
            else:
                item[ver_name] = hashed
        item[hashed][ver] = src_name
    return mindroot


def get_majver(name):
    x = re.search(r"[0-9]+", name)
    if x:
        return int(name[x.start(): x.end()])
    return 0


def tnl_by_code(ctx, model, src_name, src_ver, tgt_ver, name):
    src_majver = get_majver(src_ver)
    tgt_majver = get_majver(tgt_ver)
    if name == "${amount}":
        if isinstance(src_name, basestring):
            src_name = float(src_name)
        if (
            src_majver < 9
            and tgt_majver >= 9
            and src_name in (0.04, 0.05, 0.1, 0.21, 0.22)
        ):
            name = src_name * 100
        elif src_majver >= 9 and tgt_majver < 9 and src_name in (4, 5, 10, 21, 22):
            name = src_name / 100
        else:
            name = src_name
    elif name == "${coa}":
        if tgt_ver.split(".")[0].isdigit() and (
            src_ver.startswith("librerp")
            or src_ver.startswith("powerp")
            or src_ver.startswith("zero")
        ):
            name = {
                "121100": "1101",
                "121330": "1106",
                "121500": "1108",
                "122100": "1111",
                "122330": "1116",
                "123340": "1207",
                "122500": "1118",
                "123380": "1205",
                "123390": "1206",
                "124100": "1211",
                "124200": "1212",
                "124300": "1214",
                "124340": "1217",
                "124380": "1215",
                "124390": "1216",
                "152100": "1501",
                "152150": "1509",
                "152200": "1505",
                "152220": "1506",
                "153010": "1601",
                "153020": "1605",
                "153030": "1613",
                "153050": "1611",
                "153060": "1612",
                "153070": "1602",
                "153110": "1609",
                "190110": "1901",
                "190120": "1902",
                "211010": "2101",
                "219010": "2102",
                "240100": "2420",
                "246000": "2410",
                "250100": "2501",
                "260010": "2601",
                "260050": "2612",
                "260060": "2611",
                "260110": "2602",
                "290010": "2701",
                "290020": "2702",
                "490010": "2801",
                "490020": "2802",
                "490050": "2950",
                "510000": "3112",
                "510100": "3101",
                "510200": "3202",
                "512000": "3103",
                "550530": "3210",
                "610100": "4101",
                "610110": "4102",
                "621120": "4213",
                "621200": "4105",
                "621340": "4201",
                "623100": "4209",
                "623460": "4204",
                "623580": "4203",
                "625010": "4216",
                "625020": "4206",
                "625060": "4205",
                "625100": "4202",
                "630100": "4301",
                "640910": "4401",
                "731140": "5203",
            }.get(src_name, src_name)
        elif src_ver.split(".")[0].isdigit() and (
            tgt_ver.startswith("librerp")
            or tgt_ver.startswith("powerp")
            or tgt_ver.startswith("zero")
        ):
            name = {
                "1101": "121100",
                "1106": "121330",
                "1108": "121500",
                "1111": "122100",
                "1116": "122330",
                "1118": "122500",
                "1205": "123380",
                "1206": "123390",
                "1207": "123340",
                "1211": "124100",
                "1212": "124200",
                "1214": "124300",
                "1215": "124380",
                "1216": "124390",
                "1217": "124340",
                "1501": "152100",
                "1505": "152200",
                "1506": "152220",
                "1509": "152150",
                "1601": "153010",
                "1602": "153070",
                "1605": "153020",
                "1609": "153110",
                "1611": "153050",
                "1612": "153060",
                "1613": "153030",
                "1901": "190110",
                "1902": "190120",
                "2101": "211010",
                "2102": "219010",
                "2410": "246000",
                "2420": "240100",
                "2501": "250100",
                "2601": "260010",
                "2602": "260110",
                "2611": "260050",
                "2612": "260060",
                "2701": "290010",
                "2702": "290020",
                "2801": "490010",
                "2802": "490020",
                "2950": "490050",
                "3101": "510100",
                "3103": "512000",
                "3112": "510000",
                "3202": "510200",
                "3210": "550530",
                "4101": "610100",
                "4102": "610110",
                "4105": "621200",
                "4201": "621340",
                "4202": "625100",
                "4203": "623580",
                "4204": "623460",
                "4205": "625060",
                "4206": "625020",
                "4209": "623100",
                "4213": "621120",
                "4216": "625010",
                "4301": "630100",
                "4401": "640910",
                "5203": "731140",
            }.get(src_name, src_name)
        else:
            name = src_name
    elif name == "${tax}":
        if tgt_ver.split(".")[0].isdigit() and (
            src_ver.startswith("librerp")
            or src_ver.startswith("powerp")
            or src_ver.startswith("zero")
        ):
            name = {"a15a": "00art15a", "a15v": "00art15v"}.get(src_name, src_name)
        elif src_ver.split(".")[0].isdigit() and (
            tgt_ver.startswith("librerp")
            or tgt_ver.startswith("powerp")
            or tgt_ver.startswith("zero")
        ):
            name = {"00art15a": "a15a", "00art15v": "a15v"}.get(src_name, src_name)
        else:
            name = src_name
    return name


def previous_ver_name(ver_name, orig_name):
    x = re.search("[0-9]+", ver_name)
    if x:
        version = int(ver_name[x.start(): x.end()]) - 1
        if version < 6:
            if re.match(r"[0-9]+\.[0-9]", ver_name):
                ver_name = ""
            else:
                ver_name = orig_name if orig_name else ""
            x = re.search("[0-9]+", ver_name)
            if x:
                version = int(ver_name[x.start(): x.end()])
                ver_name = "%d.0" % version
                if "~" in orig_name:
                    ver_name = "%s~%s" % (ver_name, orig_name.split("~", 1)[-1])
            else:
                ver_name = ""
        else:
            ver_name = "%s%d%s" % (
                ver_name[0: x.start()],
                version,
                ver_name[x.end():],
            )
        if "6.0" in ver_name:
            ver_name = ver_name.replace("6.0", "6.1")
    return ver_name


def translate_from_to(
    ctx, model, src_name, src_ver, tgt_ver, ttype=False, fld_name=False, type=None
):
    """Translate symbol <src_name> from <src_ver> to <tgt_ver> of Odoo.
    If ttype not supplied, translation is applied for 'name' and 'field' ttypes
    If ttype is in ('value', 'valuetnl'), the param <fld_name> must by supplied.
    Param type is deprecated. It used just for compatibility with old version
    """
    if not ttype and type:
        ttype = type
    del type
    mindroot = ctx.get("mindroot", {})
    if src_ver not in ALL_VERSIONS:
        print("Invalid source version!")
        return ""
    if tgt_ver not in ALL_VERSIONS:
        print("Invalid target version!")
        return ""
    if ttype in ("value", "valuetnl") and not fld_name:
        print("Translation of value require field name!")
        return ""
    if ttype == "valuetnl" and not src_name:
        src_name = "dummy"
    pymodel = get_pymodel(model, ttype=ttype)
    ver_name = get_ver_name(src_name, src_ver)
    name = src_name
    if ver_name and pymodel in mindroot:
        names = []
        for typ in (
            map(lambda x: x, ("name", "field")) if not ttype else [ttype, "merge"]
        ):
            if ttype == "valuetnl":
                if fld_name in mindroot[pymodel].get("value", {}):
                    names = ["1"]
                else:
                    names = [""]
                break
            elif ttype == "value":
                item = mindroot[pymodel].get(typ, {}).get(fld_name, {})
            else:
                item = mindroot[pymodel].get(typ, {})
            if isinstance(item, basestring):
                if item.startswith("${") and item.endswith("}"):
                    # fct = item[2: -1]
                    names.append(
                        tnl_by_code(ctx, model, src_name, src_ver, tgt_ver, item)
                    )
            else:
                sver = ver_name
                while sver and sver not in item:
                    sver = previous_ver_name(sver, ver_name)
                if sver:
                    hashed = item[sver]
                    if not isinstance(hashed, list):
                        hashed = [hashed]
                    for hh in hashed:
                        if hh in item:
                            tver = tgt_ver
                            while tver and tver not in item[hh]:
                                tver = previous_ver_name(tver, tgt_ver)
                            if tver and item[hh][tver] not in names:
                                names.append(item[hh][tver])
                            if (
                                typ == "merge"
                                and src_name not in names
                                and get_majver(tver) < get_majver(sver)
                            ):
                                names.insert(0, src_name)
        if names:
            if len(names) == 1:
                name = names[0]
            else:
                name = names
        if name == r"\N":
            name = None
    return name


def translate_from_sym(ctx, model, sym, tgt_ver):
    # deprecated!
    mindroot = ctx.get("mindroot", {})
    if tgt_ver not in ALL_VERSIONS:
        print("Invalid target version!")
        return ""
    pymodel = get_pymodel(model)
    name = ""
    for typ in ("name", "field"):
        if (
            pymodel in mindroot
            and typ in mindroot[pymodel]
            and sym in mindroot[pymodel][typ]
        ):
            tver = tgt_ver
            while tver and tver not in mindroot[pymodel][typ][sym]:
                tver = previous_ver_name(tver, tgt_ver)
            if tver:
                name = mindroot[pymodel][typ][sym][tver]
                break
    return name


def model_info(ctx, pymodel, tgt_ver, ttype=False, fld_name=False):
    def do_line_ver(mindroot, model, name, typ, fld_name, tgt_ver):
        if typ not in info:
            info[model][typ] = {}
        for ver in ALL_VERSIONS:
            if tgt_ver and ver != tgt_ver:
                continue
            if ver not in info:
                info[model][typ][ver] = []
            if typ == "value":
                if ver in mindroot[model][typ][fld_name][name]:
                    info[model][typ][ver].append(
                        mindroot[model][typ][fld_name][name][ver]
                    )
            else:
                tver = ver
                while tver and tver not in mindroot[model][typ][name]:
                    tver = previous_ver_name(tver, ver)
                if tver in mindroot[model][typ][name]:
                    info[model][typ][ver].append(mindroot[model][typ][name][tver])
        return info

    mindroot = ctx.get("mindroot", {})
    if tgt_ver and tgt_ver not in ALL_VERSIONS:
        print("Invalid source version!")
        return ""
    if ttype in ("value", "valuetnl") and not fld_name:
        print("Translation of value require field name!")
        return ""
    info = {}
    for model in mindroot.keys():
        if pymodel and model != pymodel:
            continue
        if model not in info:
            info[model] = {}
        for typ in mindroot[model]:
            if ttype and typ != ttype:
                continue
            if typ == "value":
                for fld_name in mindroot[model][typ]:
                    for name in mindroot[model][typ][fld_name]:
                        if not is_hash(name):
                            continue
                        do_line_ver(mindroot, model, name, typ, fld_name, tgt_ver)
            else:
                for name in mindroot[model][typ]:
                    if not is_hash(name):
                        continue
                    do_line_ver(mindroot, model, name, typ, False, tgt_ver)
    if pymodel:
        info = info.get(pymodel, {})
        if ttype:
            info = info.get(ttype, {})
            if tgt_ver:
                info = info.get(tgt_ver, [])
    return info


def read_stored_dict(ctx):
    if "mindroot" in ctx:
        return
    if "dict_fn" not in ctx or not ctx["dict_fn"]:
        p = os.path.dirname(__file__) or "."
        if os.path.isfile("%s/transodoo.xlsx" % p):
            ctx["dict_fn"] = "%s/transodoo.xlsx" % p
        elif os.path.isfile(os.path.join(os.path.expanduser("~"), "transodoo.xlsx")):
            ctx["dict_fn"] = os.path.join(os.path.expanduser("~"), "transodoo.xlsx")
        else:
            ctx["dict_fn"] = "transodoo.xlsx"
    mindroot = {}
    wb = load_workbook(ctx["dict_fn"])
    sheet = wb.active
    colnames = []
    for column in sheet.columns:
        colnames.append(column[0].value)
    hdr = True
    for line in sheet.rows:
        if hdr:
            hdr = False
            continue
        row = {}
        for column, cell in enumerate(line):
            row[colnames[column]] = cell.value
        mindroot = build_alias_struct(
            mindroot, row["model"], row["type"], fld_name=row["name"]
        )
        ver_names = []
        used_versions = []
        last_ver = ""
        last_ver_value = ""
        for ver in ALL_VERSIONS:
            if (
                (ver.startswith("zero") and not last_ver.startswith("zero"))
                or (ver.startswith("powerp") and not last_ver.startswith("powerp"))
                or (ver.startswith("librerp") and not last_ver.startswith("librerp"))
            ):
                last_ver_value = ""
            if ver in row and row[ver] and row[ver] != last_ver_value:
                ver_names.append(row[ver])
                used_versions.append(ver)
                last_ver_value = row[ver]
            last_ver = ver
        hashed = set_hash(row["type"], row["name"], ver_names)
        for ver in used_versions:
            mindroot = link_versioned_name(
                mindroot,
                row["model"],
                hashed,
                row["type"],
                row[ver],
                ver,
                fld_name=row["name"],
            )
    ctx["mindroot"] = mindroot


def write_stored_dict(ctx):
    def write_row(sheet, line, rowid, header, widths):
        rowid += 1
        fill = "FFFFFF" if rowid % 2 else "F0F0F0"
        if not line:
            for col, item in enumerate(header):
                font_color = "000000" if col % 2 else "1F1F1F"
                cell = sheet.cell(row=rowid, column=col + 1)
                cell.value = item
                cell.font = Font(name="arial", size=8, color=font_color)
                # cell.fill = PatternFill(bgColor=fill, fill_type="solid")
                cell.fill = PatternFill(bgColor=fill)
                widths[item] = len(item)
        else:
            for col, item in enumerate(header):
                font_color = "000000" if col % 2 else "1F1F1F"
                if item in line:
                    cell = sheet.cell(row=rowid, column=col + 1)
                    cell.value = line[item]
                    cell.font = Font(name="arial", size=8, color=font_color)
                    # cell.fill = PatternFill(bgColor=fill, fill_type="solid")
                    cell.fill = PatternFill(bgColor=fill)
                    widths[item] = min(24, max(len(line[item]), widths[item]))
        return rowid

    # ctx['dict_fn'] = os.path.join(os.path.expanduser('~/transodoo.xlsx'))
    wb = Workbook()
    sheet = wb.active
    # for sheet in wb:
    #     break
    sheet.font = Font(name="arial", size=8)
    sheet.title = "transodoo"
    mindroot = ctx["mindroot"]
    rowid = 0
    header = ["model", "name", "type"]
    for ver in ALL_VERSIONS:
        header.append(ver)
    widths = {}
    rowid = write_row(sheet, None, rowid, header, widths)
    for model in sorted(mindroot.keys()):
        for ttype in sorted(mindroot[model].keys()):
            if ttype in ("value", "valuetnl"):
                iterate = sorted(mindroot[model][ttype].keys())
            else:
                iterate = [None]
            for name in iterate:
                if ttype in ("value", "valuetnl"):
                    items = mindroot[model][ttype][name]
                else:
                    items = mindroot[model][ttype]
                if isinstance(items, basestring):
                    iterate2 = [items]
                else:
                    iterate2 = sorted(items.keys())
                for hashed in iterate2:
                    if not is_hash(hashed):
                        continue
                    line = {"model": model, "type": ttype}
                    if ttype == "name":
                        line["name"] = hashed
                    else:
                        if ttype in ("value", "valuetnl"):
                            line["name"] = name
                    if isinstance(items, basestring):
                        for ver_name in VERSIONS:
                            line[ver_name] = hashed
                    else:
                        for ver_name in sorted(items[hashed].keys()):
                            if items[hashed][ver_name]:
                                line[ver_name] = items[hashed][ver_name]
                    rowid = write_row(sheet, line, rowid, header, widths)
    for col, item in enumerate(header):
        sheet.column_dimensions[get_column_letter(col + 1)].width = widths[item]
    bup_fn = "%s.bak" % ctx["dict_fn"]
    if os.path.isfile(bup_fn):
        os.unlink(bup_fn)
    os.rename(ctx["dict_fn"], bup_fn)
    wb.save(filename=ctx["dict_fn"])


def cvt_file(ctx):
    def write_row(sheet, line, rowid, header, widths):
        rowid += 1
        if not line:
            for col, item in enumerate(header):
                cell = sheet.cell(row=rowid, column=col + 1)
                cell.value = item
                cell.font = Font(name="arial", size=9)
                widths[item] = len(item)
        else:
            for col, item in enumerate(header):
                if item in line:
                    cell = sheet.cell(row=rowid, column=col + 1)
                    cell.value = line[item]
                    cell.font = Font(name="arial", size=9)
                    if line[item] and isinstance(line[item], basestring):
                        widths[item] = min(24, max(len(line[item]), widths[item]))
        return rowid

    if not ctx["file_2_cvt"]:
        print("Missed filename!\nuse -F switch")
        return 1
    wbi = load_workbook(ctx["file_2_cvt"])
    sheeti = wbi.active
    # for sheeti in wbi:
    #     break
    wbo = Workbook()
    sheeto = wbo.active
    # for sheeto in wbo:
    #     break
    sheeto.font = Font(name="arial", size=9)
    colnames = []
    for column in sheeti.columns:
        colnames.append(column[0].value)
    hdr = True
    rowid = 0
    widths = {}
    for line in sheeti.rows:
        if hdr:
            hdr = False
            new_colnames = []
            tnl_colname = {}
            for column in colnames:
                tnl_colname[column] = translate_from_to(
                    ctx,
                    ctx["model"],
                    column,
                    ctx["oe_from_ver"],
                    ctx["odoo_ver"],
                    ttype="name",
                )
                new_colnames.append(tnl_colname[column])
            rowid = write_row(sheeto, None, rowid, new_colnames, widths)
            continue
        row = {}
        for column, cell in enumerate(line):
            row[colnames[column]] = cell.value
        new_row = {}
        for item in row.keys():
            new_row[tnl_colname[item]] = translate_from_to(
                ctx,
                ctx["model"],
                row[item],
                ctx["oe_from_ver"],
                ctx["odoo_ver"],
                ttype="value",
                fld_name=item,
            )
        rowid = write_row(sheeto, new_row, rowid, new_colnames, widths)
    for col, item in enumerate(new_colnames):
        sheeto.column_dimensions[get_column_letter(col + 1)].width = widths[item]
    bup_fn = "%s.bak" % ctx["file_2_cvt"]
    if os.path.isfile(bup_fn):
        os.unlink(bup_fn)
    os.rename(ctx["file_2_cvt"], bup_fn)
    wbo.save(filename=ctx["file_2_cvt"])


def transodoo_list(ctx):
    if ctx["model"]:
        pymodel = get_pymodel(ctx["model"], ttype=ctx["opt_kind"])
    else:
        pymodel = ""
    info = model_info(
        ctx, pymodel, ctx["odoo_ver"], ttype=ctx["opt_kind"], fld_name=ctx["field_name"]
    )
    if ctx["model"]:
        print("Model: %s" % get_pymodel(ctx["model"]))
        if ctx["opt_kind"]:
            print("- Type: %s" % ctx["opt_kind"])
            if ctx["odoo_ver"]:
                print("  - Odoo version: %s" % ctx["odoo_ver"])
    if info:
        print(jsonlib.write(info, indent="    "))


def transodoo(ctx=None):
    if ctx["action"] == "list":
        read_stored_dict(ctx)
        transodoo_list(ctx)
    elif ctx["action"] == "translate":
        read_stored_dict(ctx)
        if ctx["oe_from_ver"]:
            print(
                natstr(
                    translate_from_to(
                        ctx,
                        ctx["model"],
                        ctx["sym"],
                        ctx["oe_from_ver"],
                        ctx["odoo_ver"],
                        ttype=ctx["opt_kind"],
                        fld_name=ctx["field_name"],
                    )
                )
            )
        else:
            print(
                natstr(
                    translate_from_sym(ctx, ctx["model"], ctx["sym"], ctx["odoo_ver"])
                )
            )
    elif ctx["action"] == "rewrite":
        read_stored_dict(ctx)
        write_stored_dict(ctx)
    elif ctx["action"] == "cvt-file":
        read_stored_dict(ctx)
        cvt_file(ctx)
    else:
        print("Invalid action!")
        return 1
    return 0


def main(cli_args=None):
    if not cli_args:
        cli_args = sys.argv[1:]
    parser = z0lib.parseoptargs(
        "Transodoo", "© 2017-2019 by SHS-AV s.r.l.", version=__version__
    )
    parser.add_argument("-h")
    parser.add_argument("-b", "--branch", action="store", dest="odoo_ver", default="")
    parser.add_argument(
        "-F", "--filename", action="store", dest="file_2_cvt", default=""
    )
    parser.add_argument(
        "-f", "--from-branch", action="store", dest="oe_from_ver", default=""
    )
    parser.add_argument(
        "-k", "--kind", action="store", dest="opt_kind", default="field"
    )
    parser.add_argument(
        "-l", "--language", action="store", dest="opt_lang", default="it_IT"
    )
    parser.add_argument("-m", "--model", action="store", dest="model")
    parser.add_argument("-N", "--field-name", action="store", dest="field_name")
    parser.add_argument("-n")
    parser.add_argument("-q")
    parser.add_argument("-s", "--symbol-value", action="store", dest="sym", default="")
    parser.add_argument("-V")
    parser.add_argument("-v")
    parser.add_argument("action", help="list,translate,rewrite")
    ctx = parser.parseoptargs(sys.argv[1:])
    if ctx["odoo_ver"]:
        if ctx["odoo_ver"] not in ALL_VERSIONS:
            print(
                "Invalid version %s!\nUse one of %s" % (ctx["odoo_ver"], ALL_VERSIONS)
            )
            sys.exit(1)
    if ctx["oe_from_ver"]:
        if ctx["oe_from_ver"] not in ALL_VERSIONS:
            print(
                "Invalid version %s!\nUse one of %s" % (ctx["odoo_ver"], ALL_VERSIONS)
            )
            sys.exit(1)
    sts = transodoo(ctx=ctx)
    return sts


if __name__ == "__main__":
    exit(main())
