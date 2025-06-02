#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Create map of Odoo modules
"""
from __future__ import print_function, unicode_literals
from builtins import input
from past.builtins import long, basestring
import os
import os.path as pth
import sys
from time import time, sleep
import argparse
import re
import collections
from babel.messages import pofile
from openpyxl import load_workbook, Workbook

try:
    from clodoo import clodoo
except ImportError:  # pragma: no cover
    import clodoo

# from python_plus import unicodes

__version__ = "2.0.22"


MODULE_SEP = "\ufffa"

TNL_TYPES = (
    "ir.actions.act_window,name",
    "ir.model,name",
    "ir.model.fields.field_description",
    "ir.module.category,name",
    "ir.module.module,description" "ir.module.module,shortdesc",
    "ir.module.module,summary",
    "ir.ui.menu,name",
    "ir.ui.view,arch_db",
)
INVALID_NAMES = [
    "build",
    "debian",
    "dist",
    "doc",
    "docs",
    "egg-info",
    "filestore",
    "history",
    "howtos",
    "images",
    "migrations",
    "readme",
    "redhat",
    "reference",
    "scripts",
    "server",
    "setup",
    "static",
    "tests",
    "tmp",
    "tools",
    "venv_odoo",
    "win32",
]
MAGIC_PUNCT = [" ", "!", "?", ".", ":", ";", ","]
ENGLISH_TERMS = (
    "the",
    "an",
    "and",
    "or",
    "you",
    "we",
    "they",
    "on",
    "up",
    "down",
    "from",
    "to",
    "with",
    "within",
    "off",
    "not",
    "yes",
    "be",
    "am",
    "is",
    "are",
    "been",
    "have",
    "has",
    "had",
    "allow",
    "order",
    "invoice",
    "bill",
    "tax",
    "account",
    "partner",
    "contact",
    "user",
    "product",
    "customer",
    "suppplier",
    "payment",
    "copy",
    "field",
    "preferred",
    "readonly",
    "read",
    "write",
    "required",
    "who",
    "which",
    "that",
    "what",
    "when",
    "how",
    "all",
    "must",
    "if",
    "view",
    "code",
    "name",
    "sales",
    "purchase",
    "purchases",
)
ITALIAN_TERMS = (
    "il",
    "lo",
    "la",
    "gli",
    "un",
    "uno",
    "una",
    "io",
    "tu",
    "te",
    "noi",
    "voi",
    "di",
    "del",
    "della",
    "da",
    "dal",
    "dalla",
    "su",
    "sulla",
    "con",
    "nel",
    "tra",
    "fra",
    "non",
    "si",
    "essere",
    "sono",
    "sei",
    "siamo",
    "siete",
    "ho",
    "hai",
    "ha",
    "abbiamo",
    "avete",
    "hanno",
    "avuto",
    "permettere",
    "permetti",
    "ordine",
    "ordini",
    "fattura",
    "fatture",
    "iva",
    "tassa",
    "conto",
    "nominativo",
    "contatto",
    "utente",
    "prodotto",
    "cliente",
    "fornitore",
    "pagamento",
    "incasso",
    "copia",
    "campo",
    "preferito" ", lettura",
    "leggere",
    "scrivere",
    "obbligatorio",
    "richiesto",
    "che",
    "chi",
    "cosa",
    "quando",
    "come",
    "tutto",
    "tutti",
    "deve",
    "se",
    "vista",
    "codice",
    "nome",
    "nominativo",
    "vendita",
    "vendite",
    "acquisto",
    "acquisti",
)
msg_time = time()


def msg_burst(text):  # pragma: no cover
    global msg_time
    t = time() - msg_time
    if t > 11:
        print(text, "\r")
        msg_time = time()


class OdooTranslation(object):
    """ """

    def __init__(self, opt_args):
        if sys.version_info[0] == 3:
            import translators as ts
        else:
            ts = None
        self.ts = ts
        self.opt_args = opt_args
        self.dict = {}
        if not opt_args.target_path:
            opt_args.target_path = os.getcwd()
        if (
            opt_args.target_path
            and opt_args.target_path.endswith(".po")
            and pth.basename(pth.dirname(opt_args.target_path)) == "i18n"
        ):
            self.opt_args.target_path = pth.dirname(pth.dirname(opt_args.target_path))
        if not self.opt_args.module_name and not opt_args.rewrite_xlsx:
            po_file = pth.join(opt_args.target_path, "i18n", self.opt_args.lang + ".po")
            if not pth.isfile(po_file):
                po_file = pth.join(
                    opt_args.target_path,
                    "i18n",
                    self.opt_args.lang.split("_")[0] + ".po",
                )
                if pth.isfile(po_file):
                    self.opt_args.module_name = pth.basename(self.opt_args.target_path)
        if not opt_args.file_xlsx and not self.opt_args.test:
            root = self.get_home_devel()
            if opt_args.dbg_template:
                dict_name = pth.join(root, "pypi", "tools", "odoo_template_tnl.xlsx")
            else:
                dict_name = pth.join(root, "venv", "bin", "odoo_template_tnl.xlsx")
            if pth.isfile(dict_name):
                self.opt_args.file_xlsx = dict_name
        elif not opt_args.file_xlsx and self.opt_args.test:
            root = os.environ["HOME"]
            if opt_args.dbg_template:
                dict_name = pth.join(root, "pypi", "tools", "odoo_template_tnl.xlsx")
            else:
                dict_name = pth.join(
                    pth.dirname(pth.dirname(__file__)),
                    "tests",
                    "data",
                    "odoo_template_tnl.xlsx",
                )
            if pth.isfile(dict_name):
                self.opt_args.file_xlsx = dict_name

        # Type classification
        # Name, Translatable, Grouped, Regex, Nolast
        self.types_decl = [
            (
                "odoo_model",
                False,
                False,
                (
                    r"^([a-z0-9]{2,}[\._][a-z0-9]{2,}([\._][a-z0-9]{2,})?"
                    r"|[0-9]+|/[-\w._]{2,}(/[-\w._]{2,})+)"
                ),
                False,
            ),
            (
                "word",
                True,
                True,
                (
                    r"([a-zA-Z]{1,2}[./]([a-zA-Z]{1,2}[./])*([a-zA-Z]{1,2})*"
                    r"|(\w|# )([-/]?\w| # |`)*)"
                ),
                False,
            ),
            ("punct", True, False, r"[.,:;!?]+", True),
            ("space", False, True, r"\s+", True),
            (
                "tag",
                False,
                False,
                (
                    r"(</?\w+[^/>]*/?>|%[a-zA-Z]|%\(\w+\)[a-zA-Z]"
                    r"|/w+(/w+)*|\$\{[^}]+\}|&\w+;)"
                ),
                False,
            ),
            ("space2", False, True, r"['’\"«»&]", True),
        ]
        for item in self.types_decl:
            if item[0] == "word":
                self.re_word = item[3]
            elif item[0] == "space":
                self.re_space = item[3]
            elif item[0] == "space2":
                self.re_space2 = item[3]
            elif item[0] == "tag":
                self.re_tag = item[3]
        self.re_to_upper = r".*[-.(]$"
        self.build_alias_dict()
        if not self.opt_args.prio_terms:
            if (
                not self.opt_args.database
                and not self.opt_args.module_name
                and self.opt_args.rewrite_xlsx
            ):
                self.opt_args.prio_terms = "P"
            else:
                self.opt_args.prio_terms = "D"
        if not self.opt_args.prio_terms.startswith(("P", "D", "Q")):
            if self.opt_args.verbose:
                print("Invalid switch --prio-terms value: use P, D or Q!")
            self.opt_args.prio_terms = "Q"
        self.d_ctr = self.p_ctr = 0
        if self.opt_args.merge:
            self.opt_args.merge = pth.expanduser(self.opt_args.merge)
            self.opt_args.rewrite_xlsx = True

    def get_home_devel(self):
        root = os.environ.get("HOME_DEVEL")  # pragma: no cover
        if not root or not pth.isdir(root):
            if pth.isdir(pth.expanduser("~/odoo/devel")):
                root = pth.expanduser("~/odoo/devel")
            elif pth.isdir(pth.expanduser("~/devel")):
                root = pth.expanduser("~/devel")
            else:
                root = os.environ.get("HOME")
        return root

    def ismodule(self, path):
        if pth.isdir(path):
            if (
                pth.isfile(pth.join(path, "__manifest__.py"))
                or pth.isfile(pth.join(path, "__openerp__.py"))
            ) and pth.isfile(pth.join(path, "__init__.py")):
                return True
        return False

    def isplural(self, term):
        return (
            term.endswith("s") and not term.endswith("%s") and not term.endswith(")s")
        ) or term.endswith("(s)")

    def isfullupper(self, term):
        return term == term.upper() and not any([x in term for x in MAGIC_PUNCT])

    def titlable(self, term):
        return term and term[1:] == term[1:].lower()

    def get_filenames(self, fqn=None):
        fqn = fqn or self.opt_args.file_xlsx
        tmp_file = bak_file = None
        if fqn:
            tmp_file = "%s.tmp" % fqn
            bak_file = "%s.bak" % fqn
        return fqn, tmp_file, bak_file

    def save_n_bak_fn(self, filename, tmp_file, bak_file):
        if pth.isfile(bak_file):
            os.unlink(bak_file)
        os.rename(filename, bak_file)
        os.rename(tmp_file, filename)

    def strip_magic(self, term):
        term = term.strip()
        while term and term[0] in MAGIC_PUNCT:
            term = term[1:]
        while term and term[-1] in MAGIC_PUNCT:
            term = term[:-1]
        return term

    def get_hash_key(self, key, ignore_case, module=None):
        kk = self.strip_magic(key)
        if ignore_case:
            kk = key if not self.titlable(key) else key.lower()
        if module:
            kk = module + MODULE_SEP + kk.split(MODULE_SEP)[-1]
        return kk

    def adjust_case(self, orig, tnxl):
        if tnxl:
            ix = 0
            while orig[ix] in MAGIC_PUNCT:
                if tnxl[ix] != orig[ix]:
                    tnxl = tnxl[:ix] + orig[ix] + tnxl[ix:]
                ix += 1
            if self.isfullupper(orig) and not self.titlable(tnxl):
                tnxl = tnxl.upper()
            elif len(tnxl[ix:]) > 1 and self.titlable(tnxl):
                if orig[ix].isupper() and tnxl[ix].islower():
                    tnxl = tnxl[:ix] + tnxl[ix].upper() + tnxl[ix + 1:]
                elif orig[ix].islower() and tnxl[ix].isupper():
                    tnxl = tnxl[:ix] + tnxl[ix].lower() + tnxl[ix + 1:]
            ix = -1
            while orig[ix] in MAGIC_PUNCT:
                try:
                    if tnxl[ix] != orig[ix]:
                        if ix == -1:
                            tnxl = tnxl + orig[ix]
                        else:
                            tnxl = tnxl[ix:] + orig[ix] + tnxl[ix + 1:]
                except IndexError:
                    break
                ix -= 1
        return tnxl

    def set_plural(self, orig, term):
        def plural_term(term, suffix, sep):
            if sep:
                term += sep + suffix
            else:
                term = term[:-1] + suffix
            return term

        sep = "/" if orig.endswith("(s)") else ""
        if term.endswith("ca") or term.endswith("ga"):
            term = plural_term(term, "he", sep)
        elif term.endswith("cia") or term.endswith("gia"):
            if term[-4] in ("a", "e", "i", "o", "u"):
                term = plural_term(term, "ie", sep)
            else:
                term = plural_term(term, "e", sep)
        elif term.endswith("a"):
            term = plural_term(term, "e", sep)
        elif term.endswith("e") or term.endswith("o"):
            term = plural_term(term, "i", sep)
        return term

    def set_plural_if(self, hash_key, orig, tnxl, adjust_case=None):
        if self.isplural(orig):
            hkey = hash_key[:-3] if hash_key.endswith("(s)") else hash_key[:-1]
            if hkey in self.dict:
                tnxl = self.adjust_case(orig, self.set_plural(orig, self.dict[hkey][1]))
            elif adjust_case:
                tnxl = self.adjust_case(orig, tnxl)
        elif adjust_case:
            tnxl = self.adjust_case(orig, tnxl)
        return tnxl

    def get_term(self, hash_key, orig, tnxl, adjust_case=None):
        tnxl = tnxl or orig
        if orig:
            if hash_key in self.dict:
                tnxl = self.adjust_case(orig, self.dict[hash_key][1])
            else:
                tnxl = self.set_plural_if(hash_key, orig, tnxl, adjust_case=adjust_case)
        return tnxl

    def store_1_item(
        self,
        hash_key,
        msg_orig,
        msg_tnxl,
        prio_terms="D",
        module=None,
        is_tag=None,
        raw=False,
    ):
        if len(msg_orig) <= 1 and not is_tag:
            return msg_orig
        if collections.Counter(msg_orig)["%"] != collections.Counter(msg_tnxl)["%"]:
            print(
                "*** Warning: no translation due different param subts: <<"
                + msg_orig[:80]
                + ">> / <<"
                + msg_tnxl[:80]
                + ">>"
            )
            return msg_orig
        if not is_tag and not raw:
            for item in self.tags:
                tnxls = self.dict[item]
                ltoken = "%s " % tnxls[0]
                rtoken = "%s " % tnxls[1]
                if msg_tnxl.startswith(ltoken):
                    msg_tnxl = msg_tnxl.replace(ltoken, rtoken, 1)
                ltoken = " %s " % tnxls[0]
                rtoken = " %s " % tnxls[1]
                msg_tnxl = msg_tnxl.replace(ltoken, rtoken)
                ltoken = " %s" % tnxls[0]
                rtoken = " %s" % tnxls[1]
                if msg_tnxl.endswith(ltoken):
                    msg_tnxl = msg_tnxl[0: -len(ltoken)] + rtoken
        hash_key = self.get_hash_key(hash_key, True, module=module)
        if hash_key and (
            hash_key not in self.dict
            or prio_terms == "P"
            or is_tag
            or (msg_tnxl and self.dict[hash_key][0] == self.dict[hash_key][1])
        ):
            self.dict[hash_key] = (
                self.strip_magic(msg_orig),
                self.strip_magic(msg_tnxl),
            )
            if is_tag and hash_key not in self.tags:
                self.tags.append(hash_key)
        return self.get_term(hash_key, msg_orig, msg_tnxl)

    def split_items(self, message):
        def append_group(tokens, hash_keys, groups, hash_groups):
            if re.search(self.re_space, groups[-1]):
                tokens.append(groups[:-1])
                tokens.append(groups[-1])
                hash_keys.append(hash_groups[:-1])
                hash_keys.append("")
            else:
                tokens.append(groups)
                hash_keys.append(hash_groups)
            groups = []
            hash_groups = []
            return tokens, hash_keys, groups, hash_groups

        tokens = []
        groups = []
        hash_keys = []
        hash_groups = []
        ix = 0
        while ix < len(message):
            for tok_type in self.types_decl:
                match = re.match(tok_type[3], message[ix:])
                if match:
                    token = message[ix: match.end() + ix]
                    if token.startswith("# "):
                        token = "N."
                        hash_key = self.get_hash_key(token, tok_type[0])
                        grouped = tok_type[2]
                        ix += 2
                    else:
                        hash_key = self.get_hash_key(token, tok_type[0])
                        grouped = tok_type[2]
                        ix += match.end()
                    break
            if not match:
                ii = len(message) - ix
                grouped = False
                for tok_type in self.types_decl:
                    match = re.search(tok_type[3], message[ix:])
                    if match:
                        ii = min(ii, match.start())
                        break
                tok_type = None
                token = message[ix: ii + ix]
                hash_key = self.get_hash_key(token, False)
                ix += ii
            if (
                re.search(self.re_space, token) or re.search(self.re_space2, token)
            ) and not groups:
                grouped = False
            if grouped:
                groups.append(token)
                hash_groups.append(hash_key)
            else:
                if groups:
                    tokens, hash_keys, groups, hash_groups = append_group(
                        tokens, hash_keys, groups, hash_groups
                    )
                tokens.append(token)
                if tok_type and tok_type[4] and ix >= len(message):
                    hash_keys.append("")
                else:
                    hash_keys.append(hash_key)
        if groups:
            tokens, hash_keys, groups, hash_groups = append_group(
                tokens, hash_keys, groups, hash_groups
            )
        return tokens, hash_keys

    def get_untransable_token(self, msg_orig):
        for tok_type in self.types_decl:
            if not tok_type[1]:
                if re.fullmatch(tok_type[3], msg_orig):
                    return msg_orig
        return False

    def check_if_transable(self, hashes_orig, hashes_tnxl):
        learn_about = (
            False
            if any([len(x) > 4 for x in hashes_orig if isinstance(x, (list, tuple))])
            or len(hashes_orig) == len(hashes_tnxl)
            else True
        )
        if hashes_orig != hashes_tnxl:
            ix = 0
            while not learn_about and ix < len(hashes_orig) and ix < len(hashes_tnxl):
                if (
                    (
                        isinstance(hashes_orig[ix], (list, tuple))
                        and not isinstance(hashes_tnxl[ix], (list, tuple))
                    )
                    or (
                        not isinstance(hashes_orig[ix], (list, tuple))
                        and isinstance(hashes_tnxl[ix], (list, tuple))
                    )
                    or (
                        isinstance(hashes_orig[ix], (list, tuple))
                        and isinstance(hashes_tnxl[ix], (list, tuple))
                        and (
                            len(hashes_orig[ix]) != len(hashes_tnxl[ix])
                            or any(
                                [
                                    (
                                        1 < len(hashes_orig[ix][x]) < 5
                                        and hashes_orig[ix][x] != hashes_tnxl[ix][x]
                                        and self.titlable(hashes_orig[ix][x])
                                    )
                                    for x in range(len(hashes_orig[ix]))
                                ]
                            )
                        )
                    )
                ):
                    learn_about = True
                ix += 1
        return learn_about

    def get_transated_tokens(self, tok_orig, module=None, adjust_case=None):
        hkey_orig = self.get_hash_key("".join(tok_orig), True, module=module)
        if hkey_orig in self.dict:
            return (
                self.get_term(
                    hkey_orig, "".join(tok_orig), None, adjust_case=adjust_case
                ),
                hkey_orig,
            )
        return False, False

    def process_transable_tokens(self, tok_orig, hash_orig, tok_tnxl, adjust_case=None):
        tok_tnxl = tok_tnxl or tok_orig
        term_tnxl = ""
        hash_key = ""
        ix = 0
        while ix < len(tok_orig) and ix < len(tok_tnxl):
            cur_tok = tok_orig[ix]
            if re.match(self.re_word, cur_tok):
                x = self.get_term(hash_orig[ix], cur_tok, None, adjust_case=adjust_case)
                if x != cur_tok:
                    term_tnxl = x + term_tnxl.lower()
                else:
                    term_tnxl = tok_tnxl[ix] + term_tnxl.lower()
                hash_key += hash_orig[ix]
            else:
                term_tnxl = tok_tnxl[ix] + term_tnxl.lower()
                hash_key += hash_orig[ix]
            ix += 1
        return term_tnxl, hash_key

    def parse_lang_terms(self, msg_orig, msg_tnxl):
        orig = re.split("\n|,|;|:| |-", msg_orig)
        tnxl = re.split("\n|,|;|:| |-", msg_tnxl)
        oe = [x.lower() in ENGLISH_TERMS for x in orig if x].count(True)
        oi = [x.lower() in ITALIAN_TERMS for x in orig if x].count(True)
        te = [x.lower() in ENGLISH_TERMS for x in tnxl if x].count(True)
        ti = [x.lower() in ITALIAN_TERMS for x in tnxl if x].count(True)
        return (
            "en" if oe > oi and oi <= 1 else "it" if oi > oe and oe <= 1 else "",
            "en" if te > ti and ti <= 1 else "it" if ti > te and te <= 1 else "",
        )

    def do_dict_item(
        self,
        msg_orig,
        msg_tnxl,
        action=None,
        prio_terms="D",
        module=None,
        is_tag=None,
    ):
        msg_orig = msg_orig or ""
        msg_tnxl = msg_tnxl or ""
        if not msg_orig:
            return msg_orig
        action = action or ("build_dict" if prio_terms == "P" else "translate")
        if self.get_untransable_token(msg_orig):
            return msg_orig
        msg_orig = msg_orig.replace("’", "'")
        msg_tnxl = msg_tnxl.replace("’", "'")
        lang_orig, lang_tnxl = self.parse_lang_terms(msg_orig, msg_tnxl)
        if not msg_tnxl or (
            (lang_orig == "it" and lang_tnxl != "it")
            or (lang_tnxl == "en" and lang_orig != "en")
        ):
            msg_tnxl = msg_orig
            lang_tnxl = lang_orig
        if action == "translate":
            for m in (module, None):
                for i in (True, False):
                    hash_key = self.get_hash_key(msg_orig, i, module=m)
                    if hash_key in self.dict:
                        return self.get_term(
                            hash_key, msg_orig, msg_tnxl, adjust_case=True
                        )

        hash_key = self.get_hash_key(msg_orig, False)
        if hash_key in self.dict:
            # Just build dict
            if msg_orig == msg_tnxl:
                return msg_orig
            msg_dict = self.dict[hash_key][1]
            if prio_terms == "Q":
                print("# Item <<<%s>>>" % msg_orig[:60])
                print("# Dictionary for <<<%s>>>" % msg_dict[:60])
                print("# PO file value  <<<%s>>>" % msg_tnxl[:60])
                lang_dict, lang_tnxl = self.parse_lang_terms(msg_dict, msg_tnxl)
                if (lang_dict == "it" and lang_tnxl != "it") or (
                    lang_tnxl == "en" and lang_dict != "en"
                ):
                    def_select = "D"
                elif (lang_tnxl == "it" and lang_dict != "it") or (
                    lang_dict == "en" and lang_tnxl != "en"
                ):
                    def_select = "P"
                elif (
                    self.p_ctr > 0 and self.d_ctr > (self.p_ctr * 2)
                ) or self.d_ctr > (self.p_ctr + 10):
                    def_select = "D"
                elif (
                    self.d_ctr > 0 and self.p_ctr > (self.d_ctr * 2)
                ) or self.p_ctr > (self.d_ctr + 10):
                    def_select = "P"
                else:
                    def_select = " "
                select = prio_terms
                while select not in ("D", "P"):
                    select = input("Which value (D, P)? %s\b" % def_select).upper()
                    if not select:
                        select = def_select
                    elif select.lower() == "d":
                        self.d_ctr += 1
                    elif select.lower() == "p":
                        self.p_ctr += 1
                if select == "P":
                    return self.store_1_item(
                        hash_key, msg_orig, msg_tnxl, prio_terms=prio_terms
                    )
            elif prio_terms == "P":
                return self.store_1_item(
                    hash_key, msg_orig, msg_tnxl, prio_terms=prio_terms
                )

        if action == "build_dict" and (
            (lang_orig != "it" and lang_tnxl == "it")
            or (lang_tnxl != "en" and lang_orig == "en")
        ):
            return self.store_1_item(
                hash_key,
                msg_orig,
                msg_tnxl,
                module=module,
            )

        tokens_orig, hashes_orig = self.split_items(msg_orig)
        tokens_tnxl, hashes_tnxl = self.split_items(msg_tnxl)
        learn_about = is_tag or self.check_if_transable(hashes_orig, hashes_tnxl)
        fullterm_orig = fullterm_tnxl = fulltermhk_orig = fulltermhk_tnxl = ""
        fullterm_2_store = False
        hash_key = ""
        hash_key_orig = "".join(
            ["".join(x) if isinstance(x, (list, tuple)) else x for x in hashes_orig]
        )
        if module:
            hash_key_orig = self.get_hash_key(hash_key_orig, False, module=module)

        tok_orig = tokens_orig.pop(0) if tokens_orig else ""
        hash_orig = hashes_orig.pop(0) if hashes_orig else ""
        tok_tnxl = tokens_tnxl.pop(0) if tokens_tnxl else ""
        adjust_case = not is_tag
        while tok_orig or tok_tnxl:
            if isinstance(tok_orig, (list, tuple)):
                term_orig = "".join(tok_orig)
                term_tnxl = "".join(tok_tnxl)
                term_hkey = "".join(hash_orig)
                lang_orig, lang_tnxl = self.parse_lang_terms(term_orig, term_tnxl)
                if lang_tnxl != "it":
                    term_tnxl, term_hkey = self.get_transated_tokens(
                        tok_orig, module=module, adjust_case=adjust_case
                    )
                    if not term_tnxl:
                        term_tnxl, term_hkey = self.process_transable_tokens(
                            tok_orig, hash_orig, tok_tnxl, adjust_case=adjust_case
                        )
                if learn_about or (
                    action == "build_dict" and term_tnxl != "".join(tok_orig)
                ):
                    self.store_1_item(
                        term_hkey,
                        term_orig,
                        term_tnxl,
                        module=module,
                        is_tag=is_tag,
                    )
                    fullterm_2_store = True

                hash_key += term_hkey
                fullterm_orig += term_orig
                fulltermhk_orig = fullterm_orig
                fullterm_tnxl += term_tnxl
                fulltermhk_tnxl = fullterm_tnxl
                if isinstance(tok_tnxl, (list, tuple)):
                    tok_tnxl = tokens_tnxl.pop(0) if tokens_tnxl else ""
                tok_orig = tokens_orig.pop(0) if tokens_orig else ""
                hash_orig = hashes_orig.pop(0) if hashes_orig else ""
            elif tok_orig:
                if re.match(self.re_to_upper, tok_orig):
                    adjust_case = True
                fullterm_orig += tok_orig
                hash_key += hash_orig
                while tok_tnxl and not isinstance(tok_tnxl, (list, tuple)):
                    fullterm_tnxl += tok_tnxl
                    tok_tnxl = tokens_tnxl.pop(0) if tokens_tnxl else ""
                if hash_orig:
                    fulltermhk_orig = fullterm_orig
                    fulltermhk_tnxl = fullterm_tnxl
                tok_orig = tokens_orig.pop(0) if tokens_orig else ""
                hash_orig = hashes_orig.pop(0) if hashes_orig else ""
            else:
                if isinstance(tok_tnxl, (list, tuple)):
                    fullterm_tnxl += "".join(tok_tnxl)
                else:
                    fullterm_tnxl += tok_tnxl
                fulltermhk_tnxl = fullterm_tnxl
                tok_tnxl = tokens_tnxl.pop(0) if tokens_tnxl else ""
                fullterm_2_store = True

        if learn_about:
            fullterm_orig = msg_orig
            fulltermhk_tnxl = fullterm_tnxl = msg_tnxl
            fullterm_2_store = True
            fulltermhk_orig = hash_key = hash_key_orig

        if self.get_untransable_token(fullterm_orig):
            return fullterm_orig

        if action == "build_dict":
            if not is_tag and (
                fullterm_orig.lower() == fullterm_tnxl.lower() or not fullterm_tnxl
            ):
                if self.ts and not re.search(self.re_tag, fullterm_orig):
                    try:
                        # Use Google translator
                        fullterm_tnxl = self.adjust_case(
                            fullterm_orig,
                            self.ts.google(
                                fullterm_orig,
                                from_language="en",
                                to_language=self.opt_args.lang[:2],
                                timeout=5,
                            ),
                        )
                        sleep(0.3)
                        return self.store_1_item(
                            hash_key,
                            fullterm_orig,
                            fullterm_tnxl,
                            prio_terms=prio_terms,
                        )
                    except BaseException as e:
                        print(
                            "*** Google translator error %s on <<<%s>>>!"
                            % (e, fullterm_orig)
                        )
                        sleep(2.0)
                        return self.store_1_item(
                            hash_key,
                            fullterm_orig,
                            fullterm_tnxl,
                            prio_terms=prio_terms,
                        )
            elif (fullterm_orig and fullterm_2_store) or self.isplural(fulltermhk_orig):
                return self.store_1_item(
                    hash_key,
                    fullterm_orig,
                    fullterm_tnxl.replace(
                        fulltermhk_tnxl,
                        self.set_plural_if(
                            hash_key,
                            fulltermhk_orig,
                            fulltermhk_tnxl,
                            adjust_case=not is_tag,
                        ),
                    ),
                    prio_terms=prio_terms,
                    module=module,
                    is_tag=is_tag,
                )
        elif fullterm_2_store or (
            fullterm_tnxl == fulltermhk_tnxl and hash_key in self.dict
        ):
            return self.get_term(
                hash_key, fullterm_orig, fullterm_tnxl, adjust_case=adjust_case
            )
        return fullterm_tnxl

    def translate_item(self, msg_orig, msg_tnxl, module=None, adjust_case=None):
        return self.do_dict_item(msg_orig, msg_tnxl, action="translate", module=module)

    def translate_pofile(self, po_fn):
        def add_po_line(potext, line):
            potext += line
            potext += "\n"
            return potext

        def try_to_sync(left_lines, right_lines, left_ix, right_ix):
            if left_ix < len(left_lines) and right_ix < len(right_lines):
                saved_ix = right_ix
                ctr = 6
                while ctr and left_lines[left_ix] != right_lines[right_ix]:
                    right_ix += 1
                    if right_ix < len(right_lines):
                        ctr -= 1
                    else:
                        ctr = 0
                        right_ix -= 1
                if left_lines[left_ix] == right_lines[right_ix]:
                    return left_ix, right_ix, True

                right_ix = saved_ix
                saved_ix = left_ix
                ctr = 6
                while ctr and left_lines[left_ix] != right_lines[right_ix]:
                    left_ix += 1
                    if left_ix < len(left_lines):
                        ctr -= 1
                    else:
                        ctr = 0
                        left_ix -= 1
                if left_lines[left_ix] == right_lines[right_ix]:
                    return left_ix, right_ix, True
                left_ix = saved_ix
            return left_ix, right_ix, False

        def find_diff(left_lines, right_lines, left_ix, right_ix):
            while left_ix < len(left_lines) and right_ix < len(right_lines):
                if left_lines[left_ix] == right_lines[right_ix] or (
                    re.match(r"^\".*\"$", left_lines[left_ix])
                    and re.match(r"^\".*\"$", right_lines[right_ix])
                ):
                    left_ix += 1
                    right_ix += 1
                    continue
                elif re.match(r"^\".*\"$", left_lines[left_ix]):
                    left_ix += 1
                    continue
                elif re.match(r"^\".*\"$", right_lines[right_ix]):
                    right_ix += 1
                    continue

                if left_lines[left_ix].startswith("#:"):
                    break

                left_ix, right_ix, equals = try_to_sync(
                    left_lines, right_lines, left_ix, right_ix
                )
                if equals:
                    continue

                if left_lines[left_ix].startswith("msgid") and right_lines[
                    right_ix
                ].startswith("msgid"):
                    right_ix += 1
                    left_ix += 1
                    continue
                if left_lines[left_ix].startswith("msgstr") and right_lines[
                    right_ix
                ].startswith("msgstr"):
                    right_ix += 1
                    left_ix += 1
                    continue
                break
            return left_ix, right_ix

        if pth.isfile(po_fn):
            fqn, tmp_file, bak_file = self.get_filenames(fqn=po_fn)
            if self.opt_args.verbose:
                print("# Writing %s" % fqn)

            module = self.opt_args.module_name
            try:
                catalog = pofile.read_po(open(po_fn, "r"))
            except BaseException as e:
                print("*** Error %s reading po file %s!" % (e, po_fn))
                return
            for message in catalog:
                if not message.id:
                    continue
                message.string = self.translate_item(
                    message.id, message.string, module=module
                )

            pofile.write_po(open(tmp_file, "wb"), catalog, include_previous=True)
            with open(fqn, "r") as fd:
                left_lines = fd.read().split("\n")
            with open(tmp_file, "r") as fd:
                right_lines = fd.read().split("\n")
            updated = False
            left_ix = right_ix = 0
            while left_ix < len(left_lines):
                left_ix, right_ix = find_diff(
                    left_lines, right_lines, left_ix, right_ix
                )
                if left_ix >= len(left_lines):
                    break
                if (
                    right_ix < len(right_lines)
                    and left_lines[left_ix].startswith("#:")
                    and (
                        not right_lines[right_ix].startswith("#:")
                        or left_lines[left_ix].split(":")[1]
                        != right_lines[right_ix].split(":")[1]
                    )
                ):
                    updated = True
                    right_lines.insert(right_ix, left_lines[left_ix])
                else:
                    saved_ix = left_ix
                    left_ix, right_ix, equals = try_to_sync(
                        left_lines, right_lines, left_ix, right_ix
                    )
                    if left_ix != saved_ix:
                        left_ix = saved_ix
                        updated = True
                        right_lines.insert(right_ix, left_lines[left_ix])
                left_ix += 1
                right_ix += 1

            LAST_TNL_NAME = "Antonio M. Vigliotti"
            LAST_TNL_MAIL = "antoniomaria.vigliotti@gmail.com"
            LAST_TEAM_NAME = "Zeroincombenze"
            LAST_TEAM_URL = "https://www.zeroincombenze.it"
            LANGUAGE = r'"Language: it_IT\n"'
            PLURALS = r'"Plural-Forms: nplurals=2; plural=(n != 1);\n"'
            lang_it_lineno = False
            plurals_lineno = False
            for lineno, line in enumerate(right_lines):
                if re.match(r"^#\s+\*", line):
                    right_lines[lineno] = r"# * %s" % module
                elif line.startswith("#. module:"):
                    right_lines[lineno] = r"#. module: %s" % module
                elif self.opt_args.odoo_branch and line.startswith(
                    '"Project-Id-Version:'
                ):
                    right_lines[lineno] = (
                        r'"Project-Id-Version: Odoo (%s)\n"' % self.opt_args.odoo_branch
                    )
                elif line.startswith('"Last-Translator:'):
                    right_lines[lineno] = r'"Last-Translator: %s <%s>\n"' % (
                        LAST_TNL_NAME,
                        LAST_TNL_MAIL,
                    )
                elif line.startswith('"Language-Team:'):
                    right_lines[lineno] = r'"Language-Team: %s (%s)\n"' % (
                        LAST_TEAM_NAME,
                        LAST_TEAM_URL,
                    )
                    if not lang_it_lineno:
                        lang_it_lineno = lineno
                    if not plurals_lineno:
                        plurals_lineno = lineno
                elif line.startswith('"Language:'):
                    right_lines[lineno] = LANGUAGE
                    lang_it_lineno = -1
                elif line.startswith('"Plural-Forms:'):
                    right_lines[lineno] = PLURALS
                    plurals_lineno = -1
                elif (
                    line.startswith("#")
                    and lineno > 0
                    and right_lines[lineno] == right_lines[lineno - 1]
                ):
                    del right_lines[lineno]
                    lineno -= 1
            if plurals_lineno and plurals_lineno >= 0:
                right_lines.insert(plurals_lineno, PLURALS)
                updated = True
            if lang_it_lineno and lang_it_lineno >= 0:
                right_lines.insert(lang_it_lineno, LANGUAGE)
                updated = True

            if updated:
                with open(tmp_file, "w") as fd:
                    fd.write("\n".join(right_lines))

            self.save_n_bak_fn(fqn, tmp_file, bak_file)

    def get_cache_fqn(self, fqn=None):
        fqn = fqn or self.opt_args.file_xlsx or "odoo_template_tnl.xlsx"
        return pth.join(
            pth.dirname(fqn) or os.path.expanduser("~/.local/share/odoo_translation"),
            "." + pth.basename(fqn).replace(".xlsx", "_cache.xlsx"),
        )

    def load_terms_from_cached_xlsx(self, fqn=None):
        fqn = self.get_cache_fqn(fqn=fqn)
        if not self.opt_args.ignore_cache and pth.isfile(fqn):
            self.load_terms_from_xlsx(fqn)

    def load_terms_from_pofile(self, po_fn, prio_terms="D"):
        if pth.isfile(po_fn):
            if self.opt_args.verbose:
                print("# Loading %s (prio=%s)" % (po_fn, prio_terms))
            try:
                catalog = pofile.read_po(open(po_fn, "r"))
            except BaseException as e:
                print("*** Error %s reading po file %s!" % (e, po_fn))
                return
            rowctr = 0
            for message in catalog:
                rowctr += 1
                if self.opt_args.verbose:
                    msg_burst(
                        "%d/%d) %-60.60s%s"
                        % (
                            rowctr,
                            len(self.dict),
                            message.id.split("\n")[0],
                            (
                                "> ..."
                                if "\n" in message.id or len(message.id) > 60
                                else ""
                            ),
                        )
                    )
                if not message.id:
                    continue
                if self.opt_args.wep_wrong_terms:
                    if ("\n" in message.string and "\n" not in message.id) or (
                        '"' in message.string and '"' not in message.id
                    ):
                        message.string = ""
                    else:
                        lang_msgid, lang_msgstr = self.parse_lang_terms(
                            message.id, message.string
                        )
                        if lang_msgstr == "en":
                            message.string = ""
                self.do_dict_item(
                    message.id,
                    message.string,
                    action="build_dict",
                    prio_terms=prio_terms,
                )
            if self.opt_args.verbose:
                print("# %d lines read from %s ..." % (rowctr, po_fn))

    def load_terms_from_xlsx(self, dict_fn, prio_terms="D"):
        if pth.isfile(dict_fn):
            if self.opt_args.verbose:
                print("# Loading terms from %s (prio=%s)" % (dict_fn, prio_terms))
            wb = load_workbook(dict_fn)
            sheet = wb.active
            colnames = []
            for ncol in sheet.columns:
                colnames.append(ncol[0].value)
            hdr = True
            rowctr = 0
            for nrow in sheet.rows:
                if hdr:
                    hdr = False
                    continue
                row = {}
                for ncol, cell in enumerate(nrow):
                    if cell.value and not isinstance(cell.value, basestring):
                        cell.value = str(cell.value)
                    row[colnames[ncol]] = (
                        cell.value.replace("\\n", "\n") if cell.value else cell.value
                    )
                if not row["msgid"]:
                    continue
                rowctr += 1
                if self.opt_args.verbose:
                    msg_burst(
                        "%d/%d) %-60.60s%s"
                        % (
                            rowctr,
                            len(self.dict),
                            row["msgid"].split("\n")[0],
                            (
                                "> ..."
                                if "\n" in row["msgid"] or len(row["msgid"]) > 60
                                else ""
                            ),
                        )
                    )
                if not row["msgid"] or not row["msgstr"]:
                    continue
                if self.opt_args.wep_wrong_terms:
                    if ("\n" in row["msgstr"] and "\n" not in row["msgid"]) or (
                        '"' in row["msgstr"] and '"' not in row["msgid"]
                    ):
                        continue
                    else:
                        lang_msgid, lang_msgstr = self.parse_lang_terms(
                            row["msgid"], row["msgstr"]
                        )
                        if lang_msgstr == "en":
                            continue
                if "hashkey" in row and row["hashkey"]:
                    if row["hashkey"] not in self.dict:
                        self.store_1_item(
                            row["hashkey"],
                            row["msgid"],
                            row["msgstr"],
                            module=row["module"],
                            raw=True,
                        )
                else:
                    self.do_dict_item(
                        row["msgid"],
                        row["msgstr"],
                        action="build_dict",
                        prio_terms=prio_terms,
                        module=row["module"],
                    )
            if self.opt_args.verbose:
                print("# %d lines read from %s ..." % (rowctr, dict_fn))

    def do_work_on_path(self, root, base, action=None):
        action = action or "translate"
        path = pth.join(root, base) if base else root
        if self.ismodule(path):
            i18n_path = pth.join(path, "i18n")
            po_fn = pth.join(i18n_path, "%s.po" % self.opt_args.lang)
            if not pth.isfile(po_fn):
                po_fn = pth.join(i18n_path, "%s.po" % self.opt_args.lang.split("_")[0])
            if not pth.isfile(po_fn):
                print("*** Module %s without translation!" % pth.basename(path))
                return
            if self.opt_args.merge:
                self.load_terms_from_xlsx(
                    self.opt_args.merge, prio_terms=self.opt_args.prio_terms
                )
            elif action == "build_dict":
                self.load_terms_from_pofile(po_fn, prio_terms=self.opt_args.prio_terms)
            elif action == "translate":
                self.translate_pofile(po_fn)
            else:
                raise IOError

    def build_alias_dict(self):
        self.tags = []
        for msg_orig, msg_tnxl, is_tag in (
            ("I", "io", True),
            ("a", "un", False),
            ("iva", "IVA", True),
            ("sepa", "SEPA", True),
            ("UE", "EU", True),
            ("iban", "IBAN", True),
            ("Ri.Ba", "RiBA", True),
            ("RI.BA", "RiBA", True),
            ("DDT", "DdT", True),
            ("ddt", "DdT", True),
        ):
            self.do_dict_item(msg_orig, msg_tnxl, action="build_dict", is_tag=is_tag)

    def build_dict(self):
        self.load_terms_from_cached_xlsx(self.opt_args.file_xlsx)
        if self.opt_args.file_xlsx:
            self.load_terms_from_xlsx(self.opt_args.file_xlsx)
        if not self.opt_args.module_name and self.opt_args.target_path:
            target_path = pth.abspath(self.opt_args.target_path)
            # Parse root and then all sub-directories
            self.do_work_on_path(target_path, None, action="build_dict")
            for root, dirs, files in os.walk(
                target_path, topdown=True, followlinks=False
            ):
                dirs[:] = [
                    d
                    for d in dirs
                    if (
                        not d.startswith(".")
                        and not d.startswith("_")
                        and not d.endswith("~")
                        and d not in INVALID_NAMES
                    )
                ]
                for base in dirs:
                    self.do_work_on_path(root, base, action="build_dict")

    def install_lang(self, lang, ctx):
        model = "res.lang"
        vals = {
            "lang": lang,
        }
        ids = clodoo.searchL8(ctx, model, [("code", "=", lang)])
        force = False
        if not ids:
            if self.opt_args.verbose:
                print("* Language %s not installed: installing it now" % lang)
            id = clodoo.createL8(ctx, "base.language.install", vals)
            clodoo.executeL8(ctx, "base.language.install", "lang_install", [id])
            if not isinstance(ctx["user"].id, (int, long)):
                if (
                    self.opt_args.odoo_branch
                    and int(self.opt_args.odoo_branch.split(".")[0]) >= 12
                ):
                    ctx["user"].id = 2
                else:
                    ctx["user"].id = 1
            clodoo.writeL8(ctx, "res.users", ctx["user"].id, {"lang": lang})
            force = True
        elif self.opt_args.verbose:
            print("# Found language %s" % lang)
        if force or not clodoo.searchL8(ctx, "ir.translation", [("lang", "=", lang)]):
            if self.opt_args.verbose:
                print("*** No term found for language %s: loading translation" % lang)
            id = clodoo.createL8(ctx, "base.update.translations", vals)
            clodoo.executeL8(ctx, "base.update.translations", "act_update", [id])

    def connect_db(self, database):
        try:
            uid, ctx = clodoo.oerp_set_env(
                confn=self.opt_args.config,
                db=database,
                oe_version=self.opt_args.odoo_branch,
            )
        except BaseException:
            print("*** No DB %s found!" % self.opt_args.database)
            ctx = None
        return ctx

    def translate_db(self):
        # ir.translation contains Odoo translation terms
        # @src: original (english) term
        # @source: evaluated field with source code information
        # @value: translated term
        # @type: may be [code,constraint,model,selection,sql_constraint]
        # @name: is environment name; value depends on type, may be:
        #   - @model: "model name,field name"
        #   - @code: "source file name" (format 'addons/MODULE_PATH')
        #   - @selection: "MODULE_PATH,field name"
        # @module: module which added term
        # @state: may be [translated, to_translate]
        # @res_id: id of term means:
        #   - type model: record id of model in name
        #   - type code: line number in source code (in name)
        # Report translations are in ir.ui.view model
        #
        def write_term(term):
            ctr_write = 0
            term_tnl = term
            if term.lang != self.opt_args.lang:
                domain = [
                    ("lang", "=", self.opt_args.lang),
                    ("src", "=", term.src),
                    ("type", "=", term.type),
                    ("name", "=", term.name),
                    ("module", "=", term.module),
                ]
                term_tnl = clodoo.browseL8(
                    ctx,
                    model_translation,
                    clodoo.searchL8(ctx, model_translation, domain),
                )
                if len(term_tnl):
                    term_tnl = term_tnl[0]
            value = self.translate_item(term.src, term.value, adjust_case=True)
            if not term_tnl:
                if value != term.value:
                    vals = {
                        "lang": self.opt_args.lang,
                        "value": value,
                        "state": "translated",
                        "res_id": term.res_id,
                    }
                    for name in ("src", "source", "type", "name", "module"):
                        vals[name] = term[name]
                    tnl2_id = clodoo.searchL8(
                        ctx,
                        model_translation,
                        [
                            ("type", "=", vals["type"]),
                            ("name", "=", vals["name"]),
                            ("lang", "=", vals["lang"]),
                            ("res_id", "=", vals["res_id"]),
                        ],
                    )
                    if not tnl2_id:
                        clodoo.createL8(ctx, model_translation, vals)
                    else:
                        clodoo.writeL8(ctx, model_translation, tnl2_id, vals)
                    ctr_write += 1
            elif value != term_tnl.value:
                try:
                    clodoo.writeL8(
                        ctx,
                        model_translation,
                        term.id,
                        {"value": value, "state": "translated"},
                    )
                except BaseException as e:
                    if (
                        not self.opt_args.module_name
                        or term.module == self.opt_args.module_name
                    ):
                        print("*** Error <%s> for term '%s'!" % (e, term.src))
                ctr_write += 1
            return ctr_write

        if not pth.isfile(self.opt_args.config):
            print("*** File %s not found!" % self.opt_args.config)
            return 1
        ctr_read = ctr_write = db_tnl = 0
        for database in self.opt_args.database.split(","):
            ctx = self.connect_db(database)
            if ctx is None:
                continue
            db_tnl += 1
            ctx["lang"] = self.opt_args.lang
            self.install_lang(self.opt_args.lang, ctx)
            model_translation = "ir.translation"
            for term in clodoo.browseL8(
                ctx,
                model_translation,
                clodoo.searchL8(
                    ctx,
                    model_translation,
                    [
                        ("lang", "in", ["en_US", self.opt_args.lang]),
                        ("type", "in", ["code", "model", "selection"]),
                    ],
                    order="src",
                ),
            ):
                if not term.src:
                    continue
                if self.opt_args.verbose:
                    msg_burst("%s ..." % term.src[:60].split("\n")[0])
                ctr_read += 1
                ctr_write += write_term(term)
        if db_tnl == 0:
            print("* No DB translated!")
            return 126
        if self.opt_args.verbose:
            print("# %d records read, %d records written" % (ctr_read, ctr_write))
        return 0

    def translate_module(self):
        module = self.opt_args.module_name
        target_path = pth.abspath(self.opt_args.target_path)
        if module == "OCB" and pth.isfile(pth.join(target_path, "odoo-bin")):
            self.do_work_on_path(target_path, None, action="translate")
        elif module == pth.basename(target_path):
            self.do_work_on_path(target_path, None, action="translate")
        else:
            for root, dirs, files in os.walk(
                target_path, topdown=True, followlinks=False
            ):
                dirs[:] = [
                    d
                    for d in dirs
                    if d not in (".git", "__to_remove", "doc", "setup", ".idea")
                ]
                for base in dirs:
                    if module == base:
                        self.do_work_on_path(root, base, action="translate")
                        break

    def list_dict(self):
        for hash_key, terms in self.dict.items():
            print("[%s]\n'%s'='%s'" % (hash_key, terms[0], terms[1]))

    def write_cache_xlsx(self, fqn=None):
        fqn = self.get_cache_fqn(fqn=fqn)
        if not os.path.isdir(os.path.dirname(fqn)):
            os.makedirs(os.path.dirname(fqn))
        self.write_xlsx(fqn=fqn, inplace=True)

    def write_xlsx(self, fqn=None, inplace=None, only_template=None):
        fqn, tmp_file, bak_file = self.get_filenames(fqn)
        if not pth.isfile(fqn):
            inplace = True
        if self.opt_args.verbose:
            print("# Writing %s" % fqn)
        if self.opt_args.test:
            return self.write_csv(fqn=fqn, inplace=inplace, only_template=only_template)
        wb = Workbook()
        sheet = wb.active
        sheet.title = "odoo_default_tnxl"
        sheet.cell(row=1, column=1, value="module")
        sheet.cell(row=1, column=2, value="msgid")
        sheet.cell(row=1, column=3, value="msgstr")
        if not only_template:
            sheet.cell(row=1, column=4, value="hashkey")
        row = 1
        for hash_key, terms in sorted(self.dict.items(), key=lambda x: x[0]):
            if terms[0] == terms[1]:
                continue
            if only_template and not re.match(r"[\w][-\w ]+[\w]$", terms[0]):
                continue
            row += 1
            if MODULE_SEP in hash_key:
                module = hash_key.split(MODULE_SEP)[0]
            else:
                module = ""
            sheet.cell(row=row, column=1, value=module)
            sheet.cell(row=row, column=2, value=terms[0])
            sheet.cell(row=row, column=3, value=terms[1])
            if not only_template:
                sheet.cell(row=row, column=4, value=hash_key)
        if inplace:
            wb.save(fqn)
        else:
            wb.save(tmp_file)
            self.save_n_bak_fn(fqn, tmp_file, bak_file)

    def write_csv(self, fqn=None, inplace=None, only_template=None):
        fqn = fqn.replace(".xlsx", ".csv")
        if self.opt_args.verbose:
            print("# Writing %s" % fqn)
        with open(fqn, "w") as fd:
            if only_template:
                fd.write("%s\t%s\t%s\n" % ("module", "msgid", "msgstr"))
            else:
                fd.write("%s\t%s\t%s\t%s\n" % ("module", "msgid", "msgstr", "hashkey"))
            for hash_key, terms in sorted(self.dict.items(), key=lambda x: x[0]):
                if terms[0] == terms[1]:
                    continue
                if only_template and not re.match(r"[\w][-\w ]+[\w]$", terms[0]):
                    continue
                if MODULE_SEP in hash_key:
                    module = hash_key.split(MODULE_SEP)[0]
                else:
                    module = ""
                if only_template:
                    fd.write(
                        "%s\t%s\t%s\n"
                        % (
                            module,
                            terms[0].replace("\n", "\x7f"),
                            self.adjust_case(terms[0], terms[1]).replace("\n", "\x7f"),
                        )
                    )
                else:
                    fd.write(
                        "%s\t%s\t%s\t%s\n"
                        % (
                            module,
                            terms[0].replace("\n", "\x7f"),
                            self.adjust_case(terms[0], terms[1]).replace("\n", "\x7f"),
                            hash_key,
                        )
                    )


def main(cli_args=None):
    cli_args = cli_args or sys.argv[1:]
    parser = argparse.ArgumentParser(
        description="Create translation file", epilog="© 2022-2025 by SHS-AV s.r.l."
    )
    parser.add_argument(
        "-B", "--debug-template", action="store_true", dest="dbg_template"
    )
    parser.add_argument(
        "-b",
        "--odoo-branch",
        dest="odoo_branch",
        default="12.0",
        help="Default Odoo version",
    )
    parser.add_argument("-C", "--ignore-cache", action="store_true")
    parser.add_argument("-c", "--config", help="Odoo configuration file")
    parser.add_argument(
        "-d",
        "--database",
        help="Database to translate",
    )
    parser.add_argument(
        "-G",
        "--git-orgs",
        help="Git organizations, comma separated - " "May be: oca or zero",
    )
    parser.add_argument("-l", "--lang", default="it_IT", help="Language")
    parser.add_argument("-M", "--merge")
    parser.add_argument("-m", "--module-name")
    parser.add_argument("-n", "--dry-run", action="store_true")
    parser.add_argument("-P", "--prio-terms", help="Priority terms: Dict,Po,inQuire")
    parser.add_argument("-p", "--target-path", help="Local directory of module")
    parser.add_argument("-q", "--quite", action="store_false")
    parser.add_argument("-T", "--test", action="store_true")
    parser.add_argument("-v", "--verbose", action="count", default=0)
    parser.add_argument("-V", "--version", action="version", version=__version__)
    parser.add_argument(
        "-W", "--rewrite-xlsx", action="store_true", help="Rewrite xlsx file"
    )
    parser.add_argument("-w", "--wep-wrong-terms", action="store_true")
    parser.add_argument("-x", "--file-xlsx", help="Default dictionary")

    odoo_tnxl = OdooTranslation(parser.parse_args(cli_args))
    sts = 0
    if odoo_tnxl.opt_args.merge:
        if not pth.isfile(odoo_tnxl.opt_args.merge):
            print("File %s to merge not found!" % odoo_tnxl.opt_args.merge)
            return 1
        if odoo_tnxl.opt_args.database:
            print("Invalid parameters: you cannot merge xlsx and translate DB!")
            return 1
        elif odoo_tnxl.opt_args.module_name:
            print("Invalid parameters: you cannot merge xlsx and translate module!")
            return 1
    if odoo_tnxl.opt_args.file_xlsx or odoo_tnxl.opt_args.target_path:
        odoo_tnxl.build_dict()
    else:
        print("Invalid parameters: please set xlsx file or Odoo path!")
        return 1
    if (
        odoo_tnxl.opt_args.database
        and odoo_tnxl.opt_args.file_xlsx
        and not odoo_tnxl.opt_args.rewrite_xlsx
    ):
        sts = odoo_tnxl.translate_db()
    elif odoo_tnxl.opt_args.module_name:
        odoo_tnxl.translate_module()
    if odoo_tnxl.opt_args.rewrite_xlsx:
        odoo_tnxl.write_xlsx(only_template=True)
    odoo_tnxl.write_cache_xlsx()
    return sts


if __name__ == "__main__":
    exit(main())

